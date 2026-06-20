r"""Shared Excel typography + book-tab (three-line) styling helpers.

Single source of truth for the visual conventions every ``*.xlsx`` writer
in :mod:`statspai.output` should obey:

- Times New Roman fonts (12pt title, 11pt header/body, 9pt italic notes).
- Book-tab borders: thick top rule, thin mid-rule under the header, thick
  bottom rule under the body. Mirrors LaTeX ``booktabs`` ``\toprule`` /
  ``\midrule`` / ``\bottomrule``.
- MultiIndex header support: top row holds merged panel labels, second
  row holds the per-column statistic labels.
- Auto column widths capped at 28 chars.

By centralizing here we keep ``sumstats``, ``tab``, ``collection``,
``paper_tables`` and ``modelsummary`` visually identical to
``regression_table`` / ``outreg2`` without copy-paste drift.
"""

from __future__ import annotations

from typing import Any, Optional, Sequence

import pandas as pd

# These imports are deferred to function bodies so importing this module
# does not require ``openpyxl`` until xlsx export actually runs.


# ---------------------------------------------------------------------------
# Style primitives
# ---------------------------------------------------------------------------

TIMES = "Times New Roman"
TITLE_PT = 12
HEADER_PT = 11
BODY_PT = 11
NOTES_PT = 9
_INVALID_SHEET_CHARS = set("[]:*?/\\")
PRINT_MARGIN_SIDE_IN = 0.50
PRINT_MARGIN_TOP_BOTTOM_IN = 0.75
PRINT_MARGIN_HEADER_FOOTER_IN = 0.30


def safe_sheet_name(name: object, existing: Sequence[str] = ()) -> str:
    """Return an Excel-safe, unique worksheet name.

    Excel sheet names are limited to 31 characters and cannot contain
    ``[]:*?/\\``.  This helper keeps multi-sheet writers from each having
    slightly different truncation and collision behavior.
    """
    raw = "" if name is None else str(name)
    cleaned = "".join("_" if ch in _INVALID_SHEET_CHARS else ch for ch in raw)
    cleaned = cleaned.strip("'").strip() or "Table"
    base = cleaned[:31]
    existing_lower = {str(x).lower() for x in existing}
    if base.lower() not in existing_lower:
        return base
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate.lower() not in existing_lower:
            return candidate
    raise ValueError("Could not create a unique Excel sheet name")


def _normalize_notes(notes: Optional[str | Sequence[str]]) -> list[str]:
    """Return note lines without treating a bare string as a sequence."""
    if notes is None:
        return []
    if isinstance(notes, str):
        return [notes] if notes else []
    return [str(note) for note in notes if note]


def _styles() -> dict[str, Any]:
    """Return a dict of pre-built style objects (lazy openpyxl import)."""
    from openpyxl.styles import Alignment, Border, Font, Side

    side_thin = Side(style="thin", color="000000")
    side_med = Side(style="medium", color="000000")

    return {
        "Font": Font,
        "Alignment": Alignment,
        "Border": Border,
        "Side": Side,
        "side_thin": side_thin,
        "side_med": side_med,
        "title_font": Font(bold=True, name=TIMES, size=TITLE_PT),
        "header_font": Font(bold=True, name=TIMES, size=HEADER_PT),
        "body_font": Font(name=TIMES, size=BODY_PT),
        "notes_font": Font(italic=True, name=TIMES, size=NOTES_PT),
        "center": Alignment(horizontal="center"),
        "left": Alignment(horizontal="left"),
        "border_top_thick": Border(top=side_med),
        "border_bot_thick": Border(bottom=side_med),
        "border_top_thin": Border(top=side_thin),
        "border_bot_thin": Border(bottom=side_thin),
    }


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------


def write_title(ws: Any, row: int, n_cols: int, title: str) -> int:
    """Write a centered, bold, merged title row. Returns next row to write."""
    if not title:
        return row
    s = _styles()
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = s["title_font"]
    cell.alignment = s["center"]
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    return row + 1


def write_header(
    ws: Any,
    row: int,
    df: pd.DataFrame,
    *,
    index_label: str = "",
) -> tuple[int, int]:
    """Write column headers (single- or multi-level), merging panel labels.

    Returns ``(panel_top_row, stat_row)`` — the row where the top-level
    panel labels live and the row where the per-column statistic labels
    live. For single-level headers the two rows coincide.
    """
    s = _styles()
    is_multi = isinstance(df.columns, pd.MultiIndex) and df.columns.nlevels > 1

    if is_multi:
        # ── Top: merged panel labels
        panel_top = row
        col_idx = 2
        # Use ordered unique panel labels (preserve column order)
        seen: list[object] = []
        for c in df.columns:
            if c[0] not in seen:
                seen.append(c[0])
        for panel in seen:
            n_sub = sum(1 for c in df.columns if c[0] == panel)
            cell = ws.cell(row=row, column=col_idx, value=str(panel))
            cell.font = s["header_font"]
            cell.alignment = s["center"]
            if n_sub > 1:
                ws.merge_cells(
                    start_row=row,
                    start_column=col_idx,
                    end_row=row,
                    end_column=col_idx + n_sub - 1,
                )
            col_idx += n_sub
        row += 1
        # ── Sub-row: stat names
        c0 = ws.cell(row=row, column=1, value=index_label)
        c0.font = s["header_font"]
        c0.alignment = s["left"]
        for j, col in enumerate(df.columns, 2):
            label = col[1] if isinstance(col, tuple) else col
            c = ws.cell(row=row, column=j, value=str(label))
            c.font = s["header_font"]
            c.alignment = s["center"]
        return panel_top, row

    # Single-level header
    c0 = ws.cell(row=row, column=1, value=index_label)
    c0.font = s["header_font"]
    c0.alignment = s["left"]
    for j, col in enumerate(df.columns, 2):
        c = ws.cell(row=row, column=j, value=str(col))
        c.font = s["header_font"]
        c.alignment = s["center"]
    return row, row


def write_body(
    ws: Any,
    row: int,
    df: pd.DataFrame,
    *,
    align_first_col: str = "left",
    align_data_cols: str = "center",
) -> tuple[int, int]:
    """Write the data body. Returns ``(body_top, body_bot)`` row indices."""
    s = _styles()
    align_map = {"left": s["left"], "center": s["center"]}
    body_top = row
    for idx, data_row in df.iterrows():
        c0 = ws.cell(row=row, column=1, value=str(idx))
        c0.font = s["body_font"]
        c0.alignment = align_map.get(align_first_col, s["left"])
        for j, val in enumerate(data_row, 2):
            value = "" if pd.isna(val) else str(val)
            cell = ws.cell(row=row, column=j, value=value)
            cell.font = s["body_font"]
            cell.alignment = align_map.get(align_data_cols, s["center"])
        row += 1
    return body_top, row - 1


def apply_booktab_borders(
    ws: Any,
    *,
    header_top_row: int,
    header_bot_row: int,
    body_top_row: int,
    body_bot_row: int,
    n_cols: int,
) -> None:
    r"""Stamp the three book-tab rules onto the table.

    - Thick rule above the top header row (``\toprule``).
    - Thin rule between header and body (``\midrule``).
    - Thick rule below the last body row (``\bottomrule``).

    Uses ``top=`` / ``bottom=`` borders so they layer cleanly without
    fighting any per-cell borders the caller may have set.
    """
    from openpyxl.styles import Border

    s = _styles()
    side_thin = s["side_thin"]
    side_med = s["side_med"]

    for col in range(1, n_cols + 1):
        # Top thick rule
        cell = ws.cell(row=header_top_row, column=col)
        cell.border = Border(
            top=side_med,
            bottom=cell.border.bottom,
            left=cell.border.left,
            right=cell.border.right,
        )
        # Mid thin rule = bottom of header_bot_row
        cell = ws.cell(row=header_bot_row, column=col)
        cell.border = Border(
            top=cell.border.top,
            bottom=side_thin,
            left=cell.border.left,
            right=cell.border.right,
        )
        # Bottom thick rule
        cell = ws.cell(row=body_bot_row, column=col)
        cell.border = Border(
            top=cell.border.top,
            bottom=side_med,
            left=cell.border.left,
            right=cell.border.right,
        )


def write_notes(
    ws: Any,
    row: int,
    notes: str | Sequence[str],
    n_cols: int = 1,
) -> int:
    """Write italic note rows at the bottom. Returns next row."""
    note_lines = _normalize_notes(notes)
    if not note_lines:
        return row
    s = _styles()
    for note in note_lines:
        cell = ws.cell(row=row, column=1, value=str(note))
        cell.font = s["notes_font"]
        cell.alignment = s["left"]
        if n_cols > 1:
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=n_cols
            )
        row += 1
    return row


def autofit_columns(
    ws: Any,
    n_cols: int,
    max_width: int = 28,
    min_width: int = 8,
) -> None:
    """Pick column widths from cell contents, clipped to ``[min, max]``."""
    from openpyxl.utils import get_column_letter

    for col in range(1, n_cols + 1):
        max_len = 0
        for row_cells in ws.iter_rows(
            min_col=col,
            max_col=col,
            values_only=True,
        ):
            v = row_cells[0]
            if v is not None:
                max_len = max(max_len, len(str(v)))
        width = max(min_width, min(max_len + 3, max_width))
        ws.column_dimensions[get_column_letter(col)].width = width


def apply_publication_sheet_defaults(
    ws: Any,
    *,
    header_top_row: int,
    header_bot_row: int,
    final_row: int,
    n_cols: int,
) -> None:
    """Apply workbook-view and print defaults for publication tables.

    These settings make exported workbooks usable as source files, not just
    data dumps: no distracting gridlines, stable frozen headers, repeatable
    header rows in print/PDF, fit-to-page width, and journal-friendly margins.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    n_cols = max(int(n_cols), 1)
    header_top_row = max(int(header_top_row), 1)
    header_bot_row = max(int(header_bot_row), header_top_row)
    final_row = max(int(final_row), header_bot_row + 1)

    ws.sheet_view.showGridLines = False

    freeze_col = 2 if n_cols > 1 else 1
    ws.freeze_panes = ws.cell(
        row=header_bot_row + 1,
        column=freeze_col,
    ).coordinate

    ws.print_title_rows = f"{header_top_row}:{header_bot_row}"
    last_col = get_column_letter(n_cols)
    ws.print_area = f"A1:{last_col}{final_row}"

    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape" if n_cols > 6 else "portrait"

    ws.page_margins.left = PRINT_MARGIN_SIDE_IN
    ws.page_margins.right = PRINT_MARGIN_SIDE_IN
    ws.page_margins.top = PRINT_MARGIN_TOP_BOTTOM_IN
    ws.page_margins.bottom = PRINT_MARGIN_TOP_BOTTOM_IN
    ws.page_margins.header = PRINT_MARGIN_HEADER_FOOTER_IN
    ws.page_margins.footer = PRINT_MARGIN_HEADER_FOOTER_IN


# ---------------------------------------------------------------------------
# High-level convenience: render a DataFrame as a book-tab xlsx
# ---------------------------------------------------------------------------


def render_dataframe_to_xlsx(
    df: pd.DataFrame,
    filename: str,
    *,
    title: Optional[str] = None,
    notes: Optional[str | Sequence[str]] = None,
    sheet_name: str = "Table",
    index_label: str = "",
) -> None:
    """One-shot: title + book-tab table + italic notes, all in Times.

    Used by ``sumstats`` / ``tab`` / ``paper_tables`` / ``collection``
    when they want the canonical academic three-line look.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(sheet_name)

    render_dataframe_to_sheet(
        ws,
        df,
        title=title,
        notes=notes,
        index_label=index_label,
    )

    wb.save(filename)


def render_dataframe_to_sheet(
    ws: Any,
    df: pd.DataFrame,
    *,
    title: Optional[str] = None,
    notes: Optional[str | Sequence[str]] = None,
    index_label: str = "",
    start_row: int = 1,
) -> int:
    """Render a DataFrame into an existing worksheet.

    Returns the next empty row after the rendered table and notes.  Multi-sheet
    writers use this so every output path shares the same Times/book-tab
    layout, column-width logic, MultiIndex header handling, and note styling.
    """

    n_cols = len(df.columns) + 1  # +1 for the row-label column

    row = start_row
    if title:
        row = write_title(ws, row, n_cols, title)
        row += 0  # title row already advanced; no blank padding row

    header_top, header_bot = write_header(ws, row, df, index_label=index_label)
    row = header_bot + 1
    if df.empty:
        s = _styles()
        body_top = row
        body_bot = row
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col, value="")
            cell.font = s["body_font"]
            cell.alignment = s["left"] if col == 1 else s["center"]
        row += 1
    else:
        body_top, body_bot = write_body(ws, row, df)
        row = body_bot + 1

    apply_booktab_borders(
        ws,
        header_top_row=header_top,
        header_bot_row=header_bot,
        body_top_row=body_top,
        body_bot_row=body_bot,
        n_cols=n_cols,
    )

    note_lines = _normalize_notes(notes)
    if note_lines:
        row = write_notes(ws, body_bot + 1, note_lines, n_cols=n_cols)

    autofit_columns(ws, n_cols)
    apply_publication_sheet_defaults(
        ws,
        header_top_row=header_top,
        header_bot_row=header_bot,
        final_row=max(row - 1, body_bot),
        n_cols=n_cols,
    )
    return row
