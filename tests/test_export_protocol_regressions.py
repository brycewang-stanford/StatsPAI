"""Regression guards for the ``ResultProtocolMixin`` export defaults.

Each test here pins a defect found by auditing the 2026-07 export-protocol
rollout (the commit that attached the mixin to ~165 additional result
classes). They are behavioural guards, not contract restatements — see
``test_result_protocol.py`` for the contract sweep itself.

1. Non-dataclass results whose state lives in private attributes behind
   public properties (``KMResult``) must export their real fields. They
   previously serialised to ``{}`` and wrote header-only .xlsx/.docx files
   while reporting success — a silent degradation (CLAUDE.md §7).
2. A result with genuinely nothing to export must *warn* rather than
   quietly produce an empty workbook.
3. ``to_markdown`` must not emit a literal newline inside a table cell,
   which truncates the GitHub-flavoured table at that row.
4. ``to_word`` must tolerate bespoke ``to_docx(filename)`` renderers that
   take no title argument.
5. Decomposition results with no DataFrame panel must still export via the
   Field/Value fallback instead of raising.
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass

import numpy as np
import openpyxl
import pandas as pd
import pytest

import statspai as sp
from statspai._result_serialize import ResultProtocolMixin


def _km_result():
    rng = np.random.default_rng(0)
    n = 300
    frame = pd.DataFrame(
        {
            "t": rng.exponential(1.0, n),
            "d": (rng.random(n) > 0.3).astype(int),
        }
    )
    return sp.kaplan_meier(data=frame, duration="t", event="d")


class TestPropertyBackedResults:
    """Private state + public properties must still serialise (defect 1)."""

    def test_km_result_to_dict_is_not_empty(self):
        result = _km_result()
        payload = result.to_dict()
        assert payload, "KMResult.to_dict() must not be empty"
        assert "median_survival" in payload
        assert "survival_table" in payload

    def test_km_result_excel_has_data_rows(self, tmp_path):
        path = tmp_path / "km.xlsx"
        _km_result().to_excel(str(path))
        rows = list(openpyxl.load_workbook(path).active.values)
        # header + at least one real field
        assert len(rows) > 1, f"header-only workbook written: {rows}"
        assert rows[0] == ("Field", "Value")

    def test_km_result_word_has_data_rows(self, tmp_path):
        docx = pytest.importorskip("docx")
        path = tmp_path / "km.docx"
        _km_result().to_word(str(path))
        table = docx.Document(str(path)).tables[0]
        assert len(table.rows) > 1, "header-only .docx table written"

    def test_km_result_markdown_has_body(self):
        md = _km_result().to_markdown()
        body = md.split("\n")[4:]
        assert any(line.startswith("|") for line in body), md


class TestEmptyExportWarns:
    """Nothing to export is a loud condition, not a silent empty file."""

    def test_hollow_result_warns(self, tmp_path):
        class Hollow(ResultProtocolMixin):
            def __init__(self):
                self._hidden = 1

        with pytest.warns(RuntimeWarning, match="no exportable fields"):
            Hollow().to_excel(str(tmp_path / "hollow.xlsx"))

    def test_populated_result_does_not_warn(self, tmp_path):
        @dataclass
        class Fine(ResultProtocolMixin):
            estimate: float = 1.0

        with warnings.catch_warnings(record=True) as log:
            warnings.simplefilter("always")
            Fine().to_excel(str(tmp_path / "fine.xlsx"))
        hits = [w for w in log if "no exportable fields" in str(w.message)]
        assert not hits, f"unexpected emptiness warning: {hits}"


class TestMarkdownCellSanitising:
    """A newline in a value must not truncate the table (defect 3)."""

    def test_newline_does_not_break_table(self):
        @dataclass
        class Weird(ResultProtocolMixin):
            note: str = "a|b\nsecond line"
            val: float = 1.5

        md = Weird().to_markdown()
        body = [
            line
            for line in md.split("\n")[2:]
            if line.strip() and not line.lstrip().startswith("|")
        ]
        assert not body, f"table broken by an unescaped newline: {md!r}"
        assert r"\|" in md, "pipe must stay escaped"


class TestToWordDelegation:
    """Bespoke ``to_docx(filename)`` must not receive a title (defect 4)."""

    def test_single_positional_to_docx_accepts_title(self, tmp_path):
        class OneArg(ResultProtocolMixin):
            def __init__(self):
                self.x = 1

            def to_docx(self, filename):
                with open(filename, "w", encoding="utf-8") as handle:
                    handle.write("ok")

        target = tmp_path / "one.docx"
        OneArg().to_word(str(target), "My Title")
        assert target.read_text(encoding="utf-8") == "ok"

    def test_two_positional_to_docx_receives_title(self, tmp_path):
        seen = {}

        class TwoArg(ResultProtocolMixin):
            def __init__(self):
                self.x = 1

            def to_docx(self, filename, title=None):
                seen["title"] = title
                with open(filename, "w", encoding="utf-8") as handle:
                    handle.write("ok")

        TwoArg().to_word(str(tmp_path / "two.docx"), "My Title")
        assert seen["title"] == "My Title"


class TestPanellessDecompositionExport:
    """Results with no DataFrame panel export a Field/Value sheet (defect 5)."""

    @staticmethod
    def _frame():
        rng = np.random.default_rng(0)
        n = 400
        return pd.DataFrame(
            {
                "y": rng.normal(10, 2, n),
                "x": rng.normal(size=n),
                "s1": rng.normal(5, 1, n),
                "s2": rng.normal(3, 1, n),
            }
        )

    def test_rifreg_to_excel(self, tmp_path):
        result = sp.rifreg("y ~ x", data=self._frame(), statistic="quantile", tau=0.5)
        path = tmp_path / "rif.xlsx"
        result.to_excel(str(path))
        rows = list(openpyxl.load_workbook(path).active.values)
        assert len(rows) > 1

    def test_source_decompose_to_excel(self, tmp_path):
        result = sp.source_decompose(data=self._frame(), sources=["s1", "s2"])
        path = tmp_path / "src.xlsx"
        result.to_excel(str(path))
        rows = list(openpyxl.load_workbook(path).active.values)
        assert len(rows) > 1
        assert rows[0] == ("Field", "Value")

    def test_bytes_branch_still_works(self, tmp_path):
        result = sp.rifreg("y ~ x", data=self._frame(), statistic="quantile", tau=0.5)
        payload = result.to_excel(None)
        assert isinstance(payload, bytes) and payload
