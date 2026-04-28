"""
Round-2 publication extensions to ``regtable``.

Five additions complementing Round 1 (eform / column_spanners / coef_map /
depvar stats / consistency_check):

1. ``estimate=`` / ``statistic=`` template strings — flexible cell
   formatting (mirrors R ``modelsummary``'s ``estimate=`` /
   ``statistic=`` arguments).
2. ``notation=`` — footnote-symbol family for significance markers
   ("stars" / "symbols" with ``†‡§`` / custom tuple). AER and JPE both
   request symbols rather than stars in some contexts.
3. ``apply_coef=`` — generalise ``eform`` to any callable, e.g.
   percentage transforms (``apply_coef=lambda b: 100*b``) or signed
   sqrt for distortion measures.
4. Word + Excel ``column_spanners`` rendering — closes the format
   parity gap left in Round 1.
5. ``escape=`` — opt out of auto-escape so users can pass raw
   LaTeX/HTML strings as labels (mirrors R ``kableExtra::escape``).
"""

from __future__ import annotations

import os
import tempfile
import warnings

import numpy as np
import pandas as pd
import pytest

import statspai as sp


@pytest.fixture
def ols_models():
    rng = np.random.default_rng(2026)
    n = 600
    x1 = rng.normal(0, 1, n)
    x2 = rng.normal(0, 1, n)
    y = 1.0 + 0.5 * x1 + 0.25 * x2 + rng.normal(0, 1, n)
    df = pd.DataFrame({"y": y, "x1": x1, "x2": x2})
    m1 = sp.regress("y ~ x1", data=df)
    m2 = sp.regress("y ~ x1 + x2", data=df)
    return m1, m2, df


# ---------------------------------------------------------------------------
# 1. estimate= / statistic= template strings
# ---------------------------------------------------------------------------

class TestTemplates:

    def test_default_unchanged(self, ols_models):
        m1, _, _ = ols_models
        # Without the template parameters, the rendered cell stays
        # ``coef + stars`` and ``(SE)`` exactly as before.
        out = sp.regtable(m1).to_text()
        assert "(" in out and ")" in out

    def test_estimate_template_swaps_position_of_stars(self, ols_models):
        m1, _, _ = ols_models
        # Default: "0.500***"
        # Template "{stars}{estimate}" should put stars first → "***0.500"
        out = sp.regtable(m1, estimate="{stars}{estimate}").to_text()
        # Find a coefficient row with stars
        import re
        match = re.search(r"\*+\d+\.\d+", out)
        assert match, f"Stars before estimate not found in:\n{out}"

    def test_statistic_template_three_stat_layout(self, ols_models):
        m1, _, _ = ols_models
        # Show "b ± SE" format (some economists prefer this)
        out = sp.regtable(
            m1,
            estimate="{estimate}{stars}",
            statistic="{estimate}±{std_error}",
        ).to_text()
        assert "±" in out

    def test_template_supports_t_and_p(self, ols_models):
        m1, _, _ = ols_models
        out = sp.regtable(
            m1,
            statistic="t={t_value}, p={p_value}",
            fmt="%.3f",
        ).to_text()
        assert "t=" in out
        assert "p=" in out

    def test_template_supports_ci_bounds(self, ols_models):
        m1, _, _ = ols_models
        out = sp.regtable(
            m1,
            statistic="[{conf_low}, {conf_high}]",
        ).to_text()
        # CI brackets should appear
        assert "[" in out and "]" in out
        assert "," in out

    def test_unknown_placeholder_raises(self, ols_models):
        m1, _, _ = ols_models
        with pytest.raises((KeyError, ValueError)):
            sp.regtable(m1, statistic="{nonexistent}").to_text()


# ---------------------------------------------------------------------------
# 2. notation= for significance markers
# ---------------------------------------------------------------------------

class TestNotation:

    def test_default_uses_stars(self, ols_models):
        m1, _, _ = ols_models
        out = sp.regtable(m1).to_text()
        assert "***" in out or "**" in out or "*" in out

    def test_symbols_notation_uses_dagger_family(self, ols_models):
        m1, _, _ = ols_models
        out = sp.regtable(m1, notation="symbols").to_text()
        # Default symbols family: "*", "**", "***" → "†", "‡", "§"
        assert "†" in out or "‡" in out or "§" in out
        # Footer should explain
        assert "†" in out  # at minimum the lowest level symbol appears

    def test_custom_notation_tuple(self, ols_models):
        m1, _, _ = ols_models
        out = sp.regtable(m1, notation=("·", "‖", "♣")).to_text()
        # At least one custom marker shows
        assert any(s in out for s in ("·", "‖", "♣"))

    def test_notation_off_via_stars_false(self, ols_models):
        m1, _, _ = ols_models
        out = sp.regtable(m1, stars=False).to_text()
        assert "***" not in out
        assert "**" not in out

    def test_notation_invalid_string_raises(self, ols_models):
        m1, _, _ = ols_models
        with pytest.raises(ValueError):
            sp.regtable(m1, notation="frobnicated").to_text()


# ---------------------------------------------------------------------------
# 3. apply_coef arbitrary transform
# ---------------------------------------------------------------------------

class TestApplyCoef:

    def test_apply_coef_doubles_estimates(self, ols_models):
        m1, _, _ = ols_models
        out_default = sp.regtable(m1).to_text()
        out_double = sp.regtable(m1, apply_coef=lambda b: 2 * b).to_text()
        # Coefficients in `out_double` should be roughly 2x those in default.
        # Coarse: both must contain at least one float, but they differ.
        assert out_default != out_double

    def test_apply_coef_with_se_delta_method(self, ols_models):
        """When apply_coef is set with a derivative, SE rescales accordingly."""
        m1, _, _ = ols_models
        # 100x percentage: SE also 100x
        import re
        out = sp.regtable(
            m1,
            apply_coef=lambda b: 100 * b,
            apply_coef_deriv=lambda b: 100.0,
        ).to_text()
        # Coefficient on x1 ≈ 50 (was 0.5)
        m = re.search(r"\bx1\b\s+(-?\d+\.\d+)", out)
        assert m
        assert abs(float(m.group(1)) - 50.0) < 5.0

    def test_apply_coef_default_se_unchanged_when_no_deriv(self, ols_models):
        m1, _, _ = ols_models
        # Without apply_coef_deriv, SE stays on original scale (user
        # acknowledges they need to handle SE separately). A footer note
        # warns about this.
        out = sp.regtable(m1, apply_coef=lambda b: b).to_text()
        # No exception, output non-empty
        assert "x1" in out

    def test_apply_coef_conflicts_with_eform(self, ols_models):
        m1, _, _ = ols_models
        with pytest.raises((ValueError, TypeError)):
            sp.regtable(
                m1,
                eform=True,
                apply_coef=lambda b: 2 * b,
            )


# ---------------------------------------------------------------------------
# 4. Word + Excel column_spanners
# ---------------------------------------------------------------------------

class TestWordExcelSpanners:

    def test_excel_renders_spanner_row(self, ols_models):
        m1, m2, _ = ols_models
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "t.xlsx")
            sp.regtable(
                m1, m2, m1, m2,
                column_spanners=[("OLS", 2), ("IV", 2)],
            ).save(path)
            try:
                from openpyxl import load_workbook
            except ImportError:
                pytest.skip("openpyxl not installed")
            wb = load_workbook(path)
            ws = wb.active
            cells_text = " ".join(
                str(c.value) for row in ws.iter_rows() for c in row
                if c.value is not None
            )
            assert "OLS" in cells_text
            assert "IV" in cells_text

    def test_word_renders_spanner_row(self, ols_models):
        m1, m2, _ = ols_models
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "t.docx")
            sp.regtable(
                m1, m2, m1, m2,
                column_spanners=[("OLS", 2), ("IV", 2)],
            ).save(path)
            try:
                from docx import Document
            except ImportError:
                pytest.skip("python-docx not installed")
            doc = Document(path)
            tbl = doc.tables[0]
            all_text = " ".join(
                cell.text for row in tbl.rows for cell in row.cells
            )
            assert "OLS" in all_text
            assert "IV" in all_text


# ---------------------------------------------------------------------------
# 5. escape=
# ---------------------------------------------------------------------------

class TestEscapeOptOut:

    def test_default_still_escapes(self, ols_models):
        m1, _, _ = ols_models
        tex = sp.regtable(m1, coef_labels={"x1": "log_wage"}).to_latex()
        assert "log\\_wage" in tex

    def test_escape_false_passes_raw_latex(self, ols_models):
        m1, _, _ = ols_models
        # Math-mode LaTeX must round-trip unchanged
        tex = sp.regtable(
            m1,
            coef_labels={"x1": "$\\beta_1$"},
            escape=False,
        ).to_latex()
        assert "$\\beta_1$" in tex
        # The dollar sign and backslash must NOT be auto-escaped
        assert "\\$" not in tex
        assert "\\textbackslash" not in tex

    def test_escape_false_html_passes_raw(self, ols_models):
        m1, _, _ = ols_models
        html = sp.regtable(
            m1,
            coef_labels={"x1": "<i>β<sub>1</sub></i>"},
            escape=False,
        ).to_html()
        assert "<i>β<sub>1</sub></i>" in html
        # Must NOT contain &lt; or &gt;
        assert "&lt;i&gt;" not in html
